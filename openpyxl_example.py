import structlog
from marshmallow import ValidationError
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Side, Border, borders, colors
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension

LOGGER = structlog.get_logger()


def get_worksheet(spreadsheet, sheet_name=None):
    wb = load_workbook(filename=spreadsheet)
    ws = wb.active
    if sheet_name is not None:
        try:
            ws = wb[sheet_name]
        except KeyError:
            LOGGER.error(f'{sheet_name} doesn\'t exist. Using active sheet instead')
    return ws


def get_table_start(ws, available_fields):
    first_column_name = available_fields[0]
    for row in range(ws.min_column, 100):
        for col in range(ws.min_row, 100):
            data_cell = ws.cell(column=col, row=row)
            if data_cell.value == first_column_name:
                return data_cell


def get_data_with(ws, available_fields, required_fields):
    cell = get_table_start(ws, available_fields)
    if cell is None:
        raise ValidationError(message={column: 'Missing column' for column in available_fields})
    headers = []
    for row in ws.iter_rows(min_col=cell.column, min_row=cell.row,
                            max_col=cell.column + len(available_fields) - 1, max_row=cell.row):
        headers = tuple(cell.value for cell in row)
    if headers != available_fields:
        missing_columns = set(available_fields).difference(headers)
        if set(missing_columns).intersection(required_fields):
            raise ValidationError(message={column: 'Missing column' for column in missing_columns})
    data = []
    for row in ws.iter_rows(min_col=cell.column, min_row=cell.row + 1,
                            max_col=cell.column + len(available_fields) - 1):
        row_data = {}
        for index, cell in enumerate(row):
            if headers[index] in available_fields:
                row_data[headers[index]] = cell.value
        data.append(row_data)
    return data


def fill_header(ws):
    header_fill = PatternFill(start_color=colors.COLOR_INDEX[22],
                              end_color=colors.COLOR_INDEX[22],
                              fill_type='solid')
    header_text = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_text


def resize_columns(ws):
    dim_holder = DimensionHolder(worksheet=ws)
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0),
                                                len(str(cell.value))))

    for col in range(ws.min_column, ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws,
                                                             min=col,
                                                             max=col,
                                                             width=(dims[get_column_letter(col)] + 1) * 1.2)

    ws.column_dimensions = dim_holder


def set_borders(ws):
    border = Border(left=Side(border_style=borders.BORDER_THIN, color=colors.BLACK),
                    right=Side(border_style=borders.BORDER_THIN, color=colors.BLACK),
                    top=Side(border_style=borders.BORDER_THIN, color=colors.BLACK),
                    bottom=Side(border_style=borders.BORDER_THIN, color=colors.BLACK))
    for row in ws.rows:
        for cell in row:
            cell.border = border

def send_xlsx(iterable, filename, fields, template_name, sheet_name='Config', lower_case=False):
    try:
        wb = load_workbook(os.path.join('src', 'mocks', 'excel_templates', template_name))
    except FileNotFoundError:
        wb = Workbook()
        for sheet_name in wb.sheetnames:
            del wb[sheet_name]
        wb.create_sheet(sheet_name)
    try:
        ws = wb[sheet_name]
    except KeyError:
        ws = wb.create_sheet(sheet_name)
    ws.delete_rows(ws.min_row, ws.max_row)
    ws.append(fields)
    for item in iterable:
        items = [item[field.lower() if lower_case else field] for field in fields]
        ws.append(items)

    fill_header(ws)
    resize_columns(ws)
    set_borders(ws)

    xlsxfile = BytesIO()
    wb.save(xlsxfile)
    xlsxfile.seek(0)
    mimetype = f'Content-Type: {Worksheet.mime_type}'
    return send_file(xlsxfile, attachment_filename=filename, as_attachment=True, mimetype=mimetype, cache_timeout=-1)
